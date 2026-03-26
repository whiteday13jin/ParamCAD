from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.models import ParsedInput


class InputParser:
    def parse_json_file(self, path: Path) -> ParsedInput:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
        return self.parse_payload(payload, source=f"json:{path.name}")

    def parse_excel_file(self, path: Path, sheet_name: str | None = None) -> ParsedInput:
        wb = load_workbook(path, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"Excel '{path}' is empty")

        payload = self._parse_rows_to_payload(rows)
        return self.parse_payload(payload, source=f"excel:{path.name}")

    def parse_payload(self, payload: dict[str, Any], source: str = "json") -> ParsedInput:
        if "template" not in payload:
            raise ValueError("Input payload must include 'template'")

        template = str(payload["template"]).strip()
        if not template:
            raise ValueError("Input 'template' cannot be empty")

        params = dict(payload.get("parameters", {}))
        for key, value in payload.items():
            if key in {"template", "parameters"}:
                continue
            params[key] = value

        return ParsedInput(template=template, parameters=params, raw_input=payload, source=source)

    @staticmethod
    def _parse_rows_to_payload(rows: list[tuple[Any, ...]]) -> dict[str, Any]:
        first_row = [cell for cell in rows[0] if cell is not None]
        lower_headers = [str(cell).strip().lower() for cell in first_row]

        if set(lower_headers) >= {"parameter", "value"}:
            param_idx = lower_headers.index("parameter")
            value_idx = lower_headers.index("value")
            payload: dict[str, Any] = {}
            for row in rows[1:]:
                if row is None or len(row) <= max(param_idx, value_idx):
                    continue
                key = row[param_idx]
                value = row[value_idx]
                if key is None:
                    continue
                payload[str(key).strip()] = value
            return payload

        # Header/value style: row0 = keys, row1 = values
        if len(rows) < 2:
            raise ValueError("Excel must have at least two rows for header/value mode")

        payload = {}
        headers = rows[0]
        values = rows[1]
        for idx, header in enumerate(headers):
            if header is None:
                continue
            payload[str(header).strip()] = values[idx] if idx < len(values) else None
        return payload
